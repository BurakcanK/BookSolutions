""" Excel-to-CSV Converter

Reads all the excel files in the current directory and outputs them as CSV files.
Creates different CSV files for multiple-sheet excel files.
"""

import csv
import openpyxl
import os
import sys

for excelFile in os.listdir('./excelSpreadsheets'):
    # skip non xlsx files, load the workbook
    if not excelFile.endswith('.xlsx'):
        continue

    wb = openpyxl.load_workbook(os.path.join('excelSpreadsheets', excelFile))

    # loop through every sheet in the workbook
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]

        srcDir = os.path.join('excelSpreadsheets', 'excelToCsvFiles')
        if not os.path.exists(srcDir):
            os.mkdir(srcDir)
        else:
            print("Directory 'excelToCsvFiles' already exists.", file=sys.stderr)
            sys.exit()

        # create the CSV filename
        fileName = excelFile[:-5] + '_' + sheetName + '.csv'

        # create the csv.writer
        with open(os.path.join(srcDir, fileName), 'w', newline='') as csvFile:
            csvWriter = csv.writer(csvFile)

            print("Writing data to ->", fileName)
            # loop through ever row in the sheet
            for rowNum in range(1, sheet.max_row + 1):
                rowData = []

                for colNum in range(1, sheet.max_column + 1):
                    rowData.append(sheet.cell(rowNum, colNum).value)

                csvWriter.writerow(rowData)
