""" Text Files to Spreadsheet

Reads the contents of several text files and inserts those contents into
spreadsheet, with one line of text per row and one file per column.

Usage:
    python3 textToSpreadsheet.py <directory>
"""

from openpyxl.styles import Font
from openpyxl import Workbook
import sys
import os

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 textToSpreadsheet.py <directory>")
    sys.exit()

# create workbook and select active sheet
wb = Workbook()
ws = wb.active

# make the first row bold
ws.row_dimensions[1].font = Font(bold=True)

fileCount = 1

# iterate over the files in the directory
for inFile in os.listdir(sys.argv[1]):
    if not inFile.endswith(".txt"):
        continue

    with open(os.path.join(sys.argv[1], inFile)) as textFile:
        # write the file names in the first row
        ws.cell(row=1, column=fileCount, value=inFile)

        # write each line of the file to the cells
        for (index, line) in enumerate(textFile.readlines()):
            ws.cell(row=index + 2, column=fileCount, value=line)
        fileCount += 1

wb.save('result.xlsx')
