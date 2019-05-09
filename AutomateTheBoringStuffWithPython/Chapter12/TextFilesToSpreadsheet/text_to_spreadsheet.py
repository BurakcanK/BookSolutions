"""Text Files to Spreadsheet

Read the contents of several text files and insert those contents into
spreadsheet, with one line of text per row and one file per column.

Usage:
    python3 text_to_spreadsheet.py <directory>
"""

import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 text_to_spreadsheet.py <directory>")
    sys.exit()

# create workbook and select active sheet
wb = Workbook()
ws = wb.active

# make the first row bold
ws.row_dimensions[1].font = Font(bold=True)

file_counter = 1

# iterate over the files in the directory
for infile in os.listdir(sys.argv[1]):
    if not infile.endswith(".txt"):
        continue

    with open(os.path.join(sys.argv[1], infile)) as text_file:
        # write the file names in the first row
        ws.cell(row=1, column=file_counter, value=infile)

        # write each line of the file to the cells
        for (index, line) in enumerate(text_file.readlines()):
            ws.cell(row=index + 2, column=file_counter, value=line.strip())
        file_counter += 1

wb.save(os.path.join(
    os.path.dirname(__file__),
    "result.xlsx"
))
