""" Spreadsheet to Text Files

Opens the spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.

Usage:
    python3 spreadsheetToText.py <spreadsheet>
"""

from openpyxl.utils import get_column_letter
import openpyxl
import os
import sys

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 spreadsheetToText.py <spreadsheet>")
    sys.exit()

# create workbook and select active sheet
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active

# create a dir for text files
try:
    os.makedirs('Texts')
except FileExistsError as e:
    print("Directory already exists, give another directory name.")
    sys.exit()

# each column is a file
for fileCount in range(1, ws.max_column + 1):
    # create a file for every column
    with open(os.path.join('Texts', 'text' + str(fileCount) + '.txt'), 'a') as textFile:
        # write the rows of the column to the file
        for row in range(len(ws[get_column_letter(fileCount)])):
            if ws.cell(row=row + 1, column=fileCount).value != None:
                textFile.write(ws.cell(row=row + 1, column=fileCount).value)
