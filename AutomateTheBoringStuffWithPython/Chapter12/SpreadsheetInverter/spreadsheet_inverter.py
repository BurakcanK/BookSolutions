"""Spreadsheet Cell Inverter

Invert the row and column of the cells in the spreadsheet. Basically it is
similar to taking the transpose of a matrix.

Usage:
    python3 spreadsheet_inverter.py <file>
"""

import numpy
import openpyxl
import sys
from openpyxl import Workbook

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 spreadsheet_inverter.py <file>")
    sys.exit()

# load the workbook, select active sheet
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active

# read the data
rows = []
for row in ws.iter_rows():
    rows.append([cell.value for cell in row])

# take the transpose
cols = numpy.transpose(rows)

# create the destination workbook
wb_dest = Workbook()
ws_dest = wb_dest.active

# write to destination
for col in cols:
    ws_dest.append(list(col))

wb_dest.save(sys.argv[1])
