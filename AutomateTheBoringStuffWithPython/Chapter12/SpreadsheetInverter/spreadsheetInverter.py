""" Spreadsheet Cell Inverter

Inverts the row and column of the cells in the spreadsheet. Basically it is
similar to taking the transpose of a matrix.

Usage:
    python3 spreadsheetInverter.py <file>
"""

from openpyxl import Workbook
import openpyxl
import numpy
import sys

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 spreadsheetInverter.py <file>")
    sys.exit()

# load the workbook, select active sheet
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active

# read the data
rows = []
for row in ws.iter_rows():
    rows.append([cell.value for cell in row])

# take the transpose
cols = numpy.transpose(rows)

# create the destination workbook
wbDest = Workbook()
wsDest = wbDest.active

# write to destination
for col in cols:
    wsDest.append(list(col))

wbDest.save(sys.argv[1])
