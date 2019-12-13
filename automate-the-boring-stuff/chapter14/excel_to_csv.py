"""Excel-to-CSV Converter

Read all the excel files in the current directory and output them as CSV files.
Create different CSV files for multiple-sheet excel files.
"""

import csv
import os
import openpyxl
import sys

EXCEL_FILES = os.path.join(
    os.path.dirname(__file__),
    "excel-spreadsheets"
)

OUTPUT_DIR = os.path.join(EXCEL_FILES, "excel-to-csv-files")
if not os.path.exists(OUTPUT_DIR):
    os.mkdir(OUTPUT_DIR)
else:
    print("Directory 'excel-to-csv-files' already exists.", file=sys.stderr)
    sys.exit()

for excel_file in os.listdir(EXCEL_FILES):
    if not excel_file.endswith(".xlsx"):
        continue

    wb = openpyxl.load_workbook(os.path.join("excel-spreadsheets", excel_file))

    # loop through every sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # create the CSV filename
        filename = excel_file[:-5] + "_" + sheet_name + ".csv"

        # create the csv.writer
        with open(os.path.join(OUTPUT_DIR, filename), "w", newline="") as csv_file:
            csv_writer = csv.writer(csv_file)

            print("Writing data to ->", filename)
            # loop through ever row in the sheet
            for row_num in range(1, sheet.max_row + 1):
                row_data = []

                for col_num in range(1, sheet.max_column + 1):
                    row_data.append(sheet.cell(row_num, col_num).value)

                csv_writer.writerow(row_data)
