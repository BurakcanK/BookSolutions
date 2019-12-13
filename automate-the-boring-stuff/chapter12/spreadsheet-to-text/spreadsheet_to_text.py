"""Spreadsheet to Text Files

Open the spreadsheet and write the cells of column A into one text file,
the cells of column B into another text file, and so on.

Usage:
    python3 spreadsheet_to_text.py <spreadsheet>
"""

import openpyxl
import os
import sys
from openpyxl.utils import get_column_letter

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 spreadsheet_to_text.py <spreadsheet>")
    sys.exit()

TEXT_DIR = os.path.join(
    os.path.dirname(__file__),
    "texts"
)

# create workbook and select active sheet
wb = openpyxl.load_workbook(sys.argv[1])
ws = wb.active

# create a dir for text files
try:
    os.makedirs(TEXT_DIR)
except FileExistsError as e:
    print("Directory already exists, give another directory name.")
    sys.exit()

# each column is a file
for file_count in range(1, ws.max_column + 1):
    # create a file for every column
    with open(os.path.join(TEXT_DIR, "text" + str(file_count) + ".txt"), "a") as text_file:
        # write the rows of the column to the file
        for row in range(len(ws[get_column_letter(file_count)])):
            if ws.cell(row=row + 1, column=file_count).value:
                text_file.write(ws.cell(row=row + 1, column=file_count).value)
