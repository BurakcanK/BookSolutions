"""Spreadsheet Cell Inverter

Invert the row and column of the cells in the spreadsheet. Basically it is
similar to taking the transpose of a matrix.

Usage:
    python3 spreadsheet_inverter.py <file>
"""

import numpy
import openpyxl
import os
import sys
from openpyxl import Workbook

# ensure usage
if len(sys.argv) != 2:
    print("Usage: python3 spreadsheet_inverter.py <file>")
    sys.exit()

file_path = os.path.join(
    os.path.dirname(__file__),
    sys.argv[1]
)

# load the workbook, select active sheet
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# read the data
rows = []
for row in ws.iter_rows():
    rows.append([cell.value for cell in row])

# take the transpose and cast to str
rows_t = numpy.transpose(rows)
rows_t = [list(i) for i in rows_t]
for i in range(len(rows_t)):
    rows_t[i] = [str(j) for j in rows_t[i]]

# create the destination workbook
wb_dest = Workbook()
ws_dest = wb_dest.active

# write to destination
for row in list(rows_t):
    ws_dest.append(row)

wb_dest.save(file_path)
